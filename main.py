import os
import sys
from openpyxl.styles import NamedStyle, Font, Border, Side
import tkinter as tk
from tkinter import filedialog, Text
import tkinter.font as font
from openpyxl.utils.dataframe import dataframe_to_rows
from tkinter.filedialog import asksaveasfile
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import NamedStyle, Font, Border, Side, numbers
import pandas as pd
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font

root = tk.Tk()
root.title('Power Meter Consumption Calculator (GSFA)')
myFont = font.Font(size=8)
rootloc = ''
dfLinks = pd.DataFrame(columns=['Column ', 'Row', 'Number', 'State', 'realColumn'])
dfF = pd.DataFrame(columns=['Hour', 'Sum Week', 'Sum Sat', 'Sum Sun', 'Wint Week', 'Wint Sat', 'Wint Sun'])
monthList = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
dfTotals = pd.DataFrame(columns=['Month', 'Off-peak', 'Standard', 'Peak', 'Max kVA', 'Year', 'netP', 'netS', 'netO'])
btn = []

# This is the font for the excel main headings
boldMain = NamedStyle(name="boldMain")
boldMain.font = Font(bold=False, size=11, name="Calibri")


# This is for the plain text


def func(name):
    print(name)


def createExcel(titleOfProject, df, fileLoc):
    wp = float(winterP.get())
    wo = float(winterS.get())
    ws = float(winterO.get())
    sp = float(summerP.get())
    so = float(summerS.get())
    ss = float(summerO.get())
    maxD = float(maxDT.get())
    ecbLevyV = float(ecbLevy.get())
    nefLevyV = float(nefLevy.get())
    networkTariff = float(netwLevy.get())
    declaredDFloat = float(declaredD.get())
    netMeterTariffV = float(netMeterTariff.get())

    monthDict = {'Jan': 'January',
                 'Feb': 'February',
                 'Mar': 'March',
                 'Apr': 'April',
                 'May': 'May',
                 'Jun': 'June',
                 'Jul': 'July',
                 'Aug': 'August',
                 'Sep': 'September',
                 'Oct': 'October',
                 'Nov': 'November',
                 'Dec': 'December'}

    lowSeason = {'Jan': 'January',
                 'Feb': 'February',
                 'Mar': 'March',
                 'Sep': 'September',
                 'Oct': 'October',
                 'Nov': 'November',
                 'Dec': 'December'}

    highSeason = {'Apr': 'April',
                  'May': 'May',
                  'Jun': 'June',
                  'Jul': 'July',
                  'Aug': 'August'}

    numberOfRowsPerMonth: int = 15  # This is how many spaces will be skipped per month in excel

    # Red - FF0000
    # Green - 4AE00D
    # Here is just where we import a dataframe
    highlight = NamedStyle(name="highlight")
    highlight.font = Font(bold=True, size=10, name="Impact")
    bdB = Side(style='double', color="000000")
    bdT = Side(style='thin', color="000000")

    wb = Workbook()
    wb.add_named_style(highlight)
    # print(wb.active.title)
    # print(wb.sheetnames)

    projectTitle = titleOfProject
    wb['Sheet'].title = "Sheet1"
    sh1 = wb.active
    # sh1['A1'] = projectTitle
    # sh1['A1'].font = Font(bold=True, name="Calibri", underline='single')
    sh1.column_dimensions['A'].width = 21

    # sh1['A3'] = "Note:\nThe readings marked in blue can be sent to CoW as the measurements taken for that month"
    # sh1['A3'].font = Font(bold=False, name="Calibri", underline='none')
    # sh1['A3'].fill = PatternFill("solid", fgColor='bdd7ee')
    # sh1.row_dimensions[3].height = 30
    # sh1.merge_cells('A3:H3')

    # Here we are going to populate a list with all the valid months
    listMonth = []
    for x, row in df.iterrows():
        if (row['Standard'] != ' 0.00') and (row['Off-peak'] != ' 0.00') and (row['Peak'] != ' 0.00'):
            listMonth.append(row['Month'])

    # Here we loop to contstruct the whole document
    for x, item in enumerate(listMonth):
        sh1.row_dimensions[(6 + x * numberOfRowsPerMonth)].height = 30
        sh1.column_dimensions['E'].width = 20
        sh1.column_dimensions['F'].width = 17

        # This creates the row with the month name
        tempV = monthDict.get(item) + ' ' + str("{: .0f}".format(float((df.loc[df['Month'] == item, 'Year']))))
        sh1.cell(column=1, row=(5 + x * numberOfRowsPerMonth), value=(tempV)).font = Font(bold=True, name='Calibri', underline='single')  # A5

        # This creates the row with the headings
        # sh1.cell(column=2, row=(6 + x * numberOfRowsPerMonth), value='Current').font = Font(bold=True, name='Calibri')  # B6
        # sh1.cell(column=3, row=(6 + x * numberOfRowsPerMonth), value='Previous').font = Font(bold=True, name='Calibri')  # C6
        sh1.cell(column=4, row=(6 + x * numberOfRowsPerMonth), value='Month').font = Font(bold=True, name='Calibri')  # D6
        sh1.cell(column=6, row=(6 + x * numberOfRowsPerMonth), value='Sub-Total').font = Font(bold=True, name='Calibri')  # F6

        # This creates the column with the headings
        sh1.cell(column=1, row=(7 + x * numberOfRowsPerMonth), value='Consumption: Peak').font = Font(bold=True, name='Calibri')  # A7
        sh1.cell(column=1, row=(8 + x * numberOfRowsPerMonth), value='Consumption: Standard').font = Font(bold=True, name='Calibri')  # A8
        sh1.cell(column=1, row=(9 + x * numberOfRowsPerMonth), value='Consumption: Off-peak').font = Font(bold=True, name='Calibri')  # A9
        sh1.cell(column=1, row=(10 + x * numberOfRowsPerMonth), value='Max Demand').font = Font(bold=True, name='Calibri')  # A10
        sh1.cell(column=1, row=(11 + x * numberOfRowsPerMonth), value='Net Metering: Peak').font = Font(bold=True, name='Calibri')  # A7
        sh1.cell(column=1, row=(12 + x * numberOfRowsPerMonth), value='Net Metering: Standard').font = Font(bold=True, name='Calibri')  # A8
        sh1.cell(column=1, row=(13 + x * numberOfRowsPerMonth), value='Net Metering: Off-peak').font = Font(bold=True, name='Calibri')  # A9
        sh1.cell(column=1, row=(14 + x * numberOfRowsPerMonth), value='ECB Levy').font = Font(bold=True, name='Calibri')  # A11
        sh1.cell(column=1, row=(15 + x * numberOfRowsPerMonth), value='NEF Levy').font = Font(bold=True, name='Calibri')  # A12
        sh1.cell(column=1, row=(16 + x * numberOfRowsPerMonth), value='Network Access Charge').font = Font(bold=True, name='Calibri')  # A13

        # This populates the monthly readings column
        total = float(df.loc[df['Month'] == item, "Peak"]) + float(df.loc[df['Month'] == item, "Standard"]) + float(df.loc[df['Month'] == item, "Off-peak"])
        sh1.cell(column=4, row=(7 + x * numberOfRowsPerMonth), value=float(df.loc[df['Month'] == item, "Peak"])).font = Font(bold=False, name='Calibri')  # D7
        sh1.cell(column=4, row=(8 + x * numberOfRowsPerMonth), value=float(df.loc[df['Month'] == item, "Standard"])).font = Font(bold=False, name='Calibri')  # D8
        sh1.cell(column=4, row=(9 + x * numberOfRowsPerMonth), value=float(df.loc[df['Month'] == item, "Off-peak"])).font = Font(bold=False, name='Calibri')  # D9
        sh1.cell(column=4, row=(10 + x * numberOfRowsPerMonth), value=float(df.loc[df['Month'] == item, "Max kVA"])).font = Font(bold=False, name='Calibri')  # D10
        sh1.cell(column=4, row=(11 + x * numberOfRowsPerMonth), value=float(df.loc[df['Month'] == item, "netP"])).font = Font(bold=False, name='Calibri')  # D7
        sh1.cell(column=4, row=(12 + x * numberOfRowsPerMonth), value=float(df.loc[df['Month'] == item, "netS"])).font = Font(bold=False, name='Calibri')  # D8
        sh1.cell(column=4, row=(13 + x * numberOfRowsPerMonth), value=float(df.loc[df['Month'] == item, "netO"])).font = Font(bold=False, name='Calibri')  # D9
        sh1.cell(column=4, row=(14 + x * numberOfRowsPerMonth), value=total).font = Font(bold=False, name='Calibri')  # D11
        sh1.cell(column=4, row=(15 + x * numberOfRowsPerMonth), value=total).font = Font(bold=False, name='Calibri')  # D12
        sh1.cell(column=4, row=(16 + x * numberOfRowsPerMonth), value=declaredDFloat).font = Font(bold=False, name='Calibri')  # D13

        # This fills in the tariffs
        sh1.cell(column=5, row=(10 + x * numberOfRowsPerMonth), value=maxD).font = Font(bold=False, name='Calibri')  # E10
        sh1.cell(column=5, row=(11 + x * numberOfRowsPerMonth), value=netMeterTariffV).font = Font(bold=False, name='Calibri')  # E11
        sh1.cell(column=5, row=(12 + x * numberOfRowsPerMonth), value=netMeterTariffV).font = Font(bold=False, name='Calibri')  # E12
        sh1.cell(column=5, row=(13 + x * numberOfRowsPerMonth), value=netMeterTariffV).font = Font(bold=False, name='Calibri')  # E13
        sh1.cell(column=5, row=(14 + x * numberOfRowsPerMonth), value=ecbLevyV).font = Font(bold=False, name='Calibri')  # E11
        sh1.cell(column=5, row=(15 + x * numberOfRowsPerMonth), value=nefLevyV).font = Font(bold=False, name='Calibri')  # E12
        sh1.cell(column=5, row=(16 + x * numberOfRowsPerMonth), value=networkTariff).font = Font(bold=False, name='Calibri')  # E13
        if lowSeason.get(item) != None:  # This will check that the month is low season
            sh1.cell(column=5, row=(7 + x * numberOfRowsPerMonth), value=sp).font = Font(bold=False, name='Calibri')  # E7
            sh1.cell(column=5, row=(8 + x * numberOfRowsPerMonth), value=ss).font = Font(bold=False, name='Calibri')  # E8
            sh1.cell(column=5, row=(9 + x * numberOfRowsPerMonth), value=so).font = Font(bold=False, name='Calibri')  # E9
            sh1.cell(column=5, row=(6 + x * numberOfRowsPerMonth), value='Tariff\n(Low Season)').font = Font(bold=True, name='Calibri')  # E6
            # This just gets the values in certain cells to be used in future calculations
            peak = sh1.cell(column=4, row=(7 + x * numberOfRowsPerMonth)).value
            standard = sh1.cell(column=4, row=(8 + x * numberOfRowsPerMonth)).value
            off = sh1.cell(column=4, row=(9 + x * numberOfRowsPerMonth)).value
            peakM = sh1.cell(column=4, row=(11 + x * numberOfRowsPerMonth)).value
            standardM = sh1.cell(column=4, row=(12 + x * numberOfRowsPerMonth)).value
            offM = sh1.cell(column=4, row=(13 + x * numberOfRowsPerMonth)).value
            tarrifP = sh1.cell(column=5, row=(7 + x * numberOfRowsPerMonth)).value
            tarrifS = sh1.cell(column=5, row=(8 + x * numberOfRowsPerMonth)).value
            tarrifO = sh1.cell(column=5, row=(9 + x * numberOfRowsPerMonth)).value
            tarrifNetMetering = sh1.cell(column=5, row=(11 + x * numberOfRowsPerMonth)).value
        else:
            sh1.cell(column=5, row=(6 + x * numberOfRowsPerMonth), value='Tariff\n(High Season)').font = Font(bold=True, name='Calibri')  # E6
            sh1.cell(column=5, row=(7 + x * numberOfRowsPerMonth), value=wp).font = Font(bold=False, name='Calibri')  # E7
            sh1.cell(column=5, row=(8 + x * numberOfRowsPerMonth), value=ws).font = Font(bold=False, name='Calibri')  # E8
            sh1.cell(column=5, row=(9 + x * numberOfRowsPerMonth), value=wo).font = Font(bold=False, name='Calibri')  # E9
            # This just gets the values in certain cells to be used in future calculations
            peak = sh1.cell(column=4, row=(7 + x * numberOfRowsPerMonth)).value
            standard = sh1.cell(column=4, row=(8 + x * numberOfRowsPerMonth)).value
            off = sh1.cell(column=4, row=(9 + x * numberOfRowsPerMonth)).value
            peakM = sh1.cell(column=4, row=(11 + x * numberOfRowsPerMonth)).value
            standardM = sh1.cell(column=4, row=(12 + x * numberOfRowsPerMonth)).value
            offM = sh1.cell(column=4, row=(13 + x * numberOfRowsPerMonth)).value
            tarrifP = sh1.cell(column=5, row=(7 + x * numberOfRowsPerMonth)).value
            tarrifS = sh1.cell(column=5, row=(8 + x * numberOfRowsPerMonth)).value
            tarrifO = sh1.cell(column=5, row=(9 + x * numberOfRowsPerMonth)).value
            tarrifNetMetering = sh1.cell(column=5, row=(11 + x * numberOfRowsPerMonth)).value
        # This will centre the text in the heading cells
        sh1.cell(column=2, row=(6 + x * numberOfRowsPerMonth)).alignment = Alignment(horizontal="center", vertical="center")
        sh1.cell(column=3, row=(6 + x * numberOfRowsPerMonth)).alignment = Alignment(horizontal="center", vertical="center")
        sh1.cell(column=4, row=(6 + x * numberOfRowsPerMonth)).alignment = Alignment(horizontal="center", vertical="center")
        sh1.cell(column=5, row=(6 + x * numberOfRowsPerMonth)).alignment = Alignment(horizontal="center", vertical="center")
        sh1.cell(column=6, row=(6 + x * numberOfRowsPerMonth)).alignment = Alignment(horizontal="center", vertical="center")

        # This gets the value for the month
        maxDfromData = sh1.cell(column=4, row=(10 + x * numberOfRowsPerMonth)).value
        networkSize = sh1.cell(column=4, row=(13 + x * numberOfRowsPerMonth)).value
        totalForLevy = total

        # This is the sub totals column
        sh1.cell(column=6, row=(7 + x * numberOfRowsPerMonth), value=tarrifP * peak).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(8 + x * numberOfRowsPerMonth), value=tarrifS * standard).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(9 + x * numberOfRowsPerMonth), value=tarrifO * off).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(10 + x * numberOfRowsPerMonth), value=maxD * maxDfromData).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(11 + x * numberOfRowsPerMonth), value=-tarrifNetMetering * peakM).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(12 + x * numberOfRowsPerMonth), value=-tarrifNetMetering * standardM).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(13 + x * numberOfRowsPerMonth), value=-tarrifNetMetering * offM).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(14 + x * numberOfRowsPerMonth), value=ecbLevyV * totalForLevy).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(15 + x * numberOfRowsPerMonth), value=nefLevyV * totalForLevy).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(16 + x * numberOfRowsPerMonth), value=declaredDFloat * networkTariff).font = Font(bold=False, name='Calibri')

        # This is where the cells in the sub-total column will be formatted as number type
        stringT = r'_("N$"* #,##0.00_);_("N$"* -#,##0.00_);_("N$"* "-"??_);_(@_)'
        sh1.cell(column=6, row=(7 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(8 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(9 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(10 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(11 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(12 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(13 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(14 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(15 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(16 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(17 + x * numberOfRowsPerMonth)).number_format = stringT
        sh1.cell(column=6, row=(18 + x * numberOfRowsPerMonth)).number_format = stringT

        # This is probably a bad way of getting and summing these values but here goes
        v7 = sh1.cell(column=6, row=(7 + x * numberOfRowsPerMonth)).value
        v8 = sh1.cell(column=6, row=(8 + x * numberOfRowsPerMonth)).value
        v9 = sh1.cell(column=6, row=(9 + x * numberOfRowsPerMonth)).value
        v10 = sh1.cell(column=6, row=(10 + x * numberOfRowsPerMonth)).value
        v11 = sh1.cell(column=6, row=(11 + x * numberOfRowsPerMonth)).value
        v12 = sh1.cell(column=6, row=(12 + x * numberOfRowsPerMonth)).value
        v13 = sh1.cell(column=6, row=(13 + x * numberOfRowsPerMonth)).value
        v14 = sh1.cell(column=6, row=(14 + x * numberOfRowsPerMonth)).value
        v15 = sh1.cell(column=6, row=(15 + x * numberOfRowsPerMonth)).value
        v16 = sh1.cell(column=6, row=(16 + x * numberOfRowsPerMonth)).value

        # This is where we sum all of the monthly rows
        sh1.cell(column=6, row=(17 + x * numberOfRowsPerMonth), value=(v7 + v8 + v9 + v10 + v11 + v12 + v13 + v14 + v15 + v16)).font = Font(bold=False, name='Calibri')
        sh1.cell(column=6, row=(18 + x * numberOfRowsPerMonth), value=(sh1.cell(column=6, row=(17 + x * numberOfRowsPerMonth)).value) * (1.15)).font = Font(bold=True, name='Calibri')

        # This is where we make needed cells that blue color
        # for count in range(7, 10):
        #     sh1.cell(column=2, row=(count + x * numberOfRowsPerMonth), value='Type_formula').font = Font(italic=False, name="Calibri")
        #     sh1.cell(column=2, row=(count + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='bdd7ee')

        # This is where the cells are colored
        sh1.cell(column=4, row=(7 + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='ff4545')  # This is peak
        sh1.cell(column=4, row=(8 + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='FFFF00')  # This is standard
        sh1.cell(column=4, row=(9 + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='00CD00')  # This is off
        sh1.cell(column=4, row=(10 + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='bdd7ee')
        sh1.cell(column=4, row=(11 + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='ff4545')  # This is peak
        sh1.cell(column=4, row=(12 + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='FFFF00')  # This is standard
        sh1.cell(column=4, row=(13 + x * numberOfRowsPerMonth)).fill = PatternFill("solid", fgColor='00CD00')  # This is off

        # This is just where some things are finalized and borders added
        sh1.cell(column=7, row=(17 + x * numberOfRowsPerMonth), value='Excl. VAT')
        sh1.cell(column=7, row=(18 + x * numberOfRowsPerMonth), value='VAT (15%)')
        # sh1.cell(column=1, row=(14 + x * numberOfRowsPerMonth), value='(assume 100kVA)')
        sh1.cell(column=5, row=(17 + x * numberOfRowsPerMonth), value='Total:').font = Font(bold=True, name="Calibri")
        sh1.cell(column=5, row=(17 + x * numberOfRowsPerMonth)).border = Border(bottom=bdB, top=bdT)
        sh1.cell(column=6, row=(17 + x * numberOfRowsPerMonth)).border = Border(bottom=bdB, top=bdT)

    wb.save(os.path.join(fileLoc.name))
    return


def changeState(btn, nbr, column, row):
    if btn[nbr]['text'] == 'S':

        btn[nbr].config(text="O", fg="White", bg="Green")
    elif btn[nbr]['text'] == 'O':
        btn[nbr].config(text="P", fg="White", bg="Red")
    elif btn[nbr]['text'] == 'P':
        btn[nbr].config(text="S", fg="Black", bg="Yellow")
    return


def selectIn():
    global rootloc
    filename = filedialog.askopenfilename(initialdir="/", title="Select File",
                                          filetypes=(("csv file ", "*.csv"), ("all files", "*.*")))
    rootloc = filename
    return


def exit():
    sys.exit()
    return


def genOutput():
    for x, item in enumerate(monthList):
        dfTotals.loc[x] = [item] + [0] + [0] + [0] + [0] + [0] + [0] + [0] + [0]

    maxVA = 0
    up = 2
    float1 = 0
    startFlag = 0
    csvFile = open(rootloc)
    dateSplit = [2019, 1, 1]
    lines = csvFile.readlines()
    for p in lines:
        p1 = p.replace('"', '')
        p2 = p1.split(',')

        if (p.find('Date') != -1) and (p.find('Total VA') != -1):
            startFlag = 1

        if startFlag == 1:
            try:
                dateI = p2.index('Date')
                totalI = p2.index('Total VA')
                startI = p2.index('Start')
                endI = p2.index('End')
                importI = p2.index('Import W')
                exportI = p2.index('Export W')
            except ValueError:
                number = -1
                for x, item in enumerate(monthList):
                    if p.find(monthList[x]) != -1:
                        number = p.find(monthList[x])

                if number != -1:
                    up = up + 1
                    if monthList[dateSplit[1] - 1] != p2[dateI].split('-')[1]:
                        maxVA = 0
                    else:
                        pass
                    dateSplit = p2[dateI].split('-')
                    dateInteger = monthList.index(dateSplit[1]) + 1
                    dateSplit[1] = dateInteger
                    dateSplit[2] = '20' + dateSplit[2]
                    startHour = int(p2[startI].split(":")[0])
                    day = datetime.datetime(int(dateSplit[2]), dateSplit[1], int(dateSplit[0])).weekday()
                    float1 += float(p2[importI])
                    # This is where we try to populate the year colum
                    dfTotals.loc[(dateSplit[1] - 1), 'Year'] = (float(dateSplit[2]))

                    if (float(p2[totalI]) > maxVA):
                        dfTotals.loc[(dateSplit[1] - 1), 'Max kVA'] = float(p2[totalI])
                        maxVA = float(p2[totalI])
                    if (dateInteger < 4) or (dateInteger > 8):
                        if (day < 5):
                            if (dfF.iloc[startHour]['Sum Week'] == 'O'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Off-peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netO'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Sum Week'] == 'S'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Standard'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netS'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Sum Week'] == 'P'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netP'] += float(p2[exportI]) / 2
                        elif (day == 5):
                            if (dfF.iloc[startHour]['Sum Sat'] == 'O'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Off-peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netO'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Sum Sat'] == 'S'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Standard'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netS'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Sum Sat'] == 'P'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netP'] += float(p2[exportI]) / 2
                        elif (day == 6):
                            if (dfF.iloc[startHour]['Sum Sun'] == 'O'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Off-peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netO'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Sum Sun'] == 'S'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Standard'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netS'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Sum Sun'] == 'P'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netP'] += float(p2[exportI]) / 2
                    elif (dateInteger > 3) and (dateInteger < 9):
                        if (day < 5):
                            if (dfF.iloc[startHour]['Wint Week'] == 'O'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Off-peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netO'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Wint Week'] == 'S'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Standard'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netS'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Wint Week'] == 'P'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netP'] += float(p2[exportI]) / 2
                        elif (day == 5):
                            if (dfF.iloc[startHour]['Wint Sat'] == 'O'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Off-peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netO'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Wint Sat'] == 'S'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Standard'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netS'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Wint Sat'] == 'P'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netP'] += float(p2[exportI]) / 2
                        elif (day == 6):
                            if (dfF.iloc[startHour]['Wint Sun'] == 'O'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Off-peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netO'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Wint Sun'] == 'S'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Standard'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netS'] += float(p2[exportI]) / 2
                            elif (dfF.iloc[startHour]['Wint Sun'] == 'P'):
                                dfTotals.loc[(dateSplit[1] - 1), 'Peak'] += float(p2[importI]) / 2
                                dfTotals.loc[(dateSplit[1] - 1), 'netP'] += float(p2[exportI]) / 2

    for x, row in dfTotals.iterrows():
        dfTotals.iloc[x]["Off-peak"] = "{: .2f}".format(dfTotals.iloc[x]["Off-peak"])
        dfTotals.iloc[x]["Peak"] = "{: .2f}".format(dfTotals.iloc[x]["Peak"])
        dfTotals.iloc[x]["Standard"] = "{: .2f}".format(dfTotals.iloc[x]["Standard"])
        dfTotals.iloc[x]["netO"] = "{: .2f}".format(dfTotals.iloc[x]["netO"])
        dfTotals.iloc[x]["netP"] = "{: .2f}".format(dfTotals.iloc[x]["netP"])
        dfTotals.iloc[x]["netS"] = "{: .2f}".format(dfTotals.iloc[x]["netS"])
        dfTotals.iloc[x]["'Max kVA'"] = "{: .2f}".format(dfTotals.iloc[x]["Max kVA"])
        dfTotals.iloc[x]["Year"] = "{: .0f}".format(dfTotals.iloc[x]["Year"])

    file = asksaveasfile(initialdir="/", title="Select output location",
                         filetypes=(("Excel File", "*.xlsx"), ("all files", "*.*")), defaultextension=".xlsx")

    dfTotalsString = dfTotals.round(1).astype(str)
    dfTotalsString['Off-peak'] += " kWh"
    dfTotalsString['Standard'] += " kWh"
    dfTotalsString['Peak'] += " kWh"
    dfTotalsString['Max kVA'] += " kVA"
    # dfTotalsString.to_csv(file, index=False)
    createExcel("Type_the_project_title", dfTotals, file)
    return


def confirmSched(btn):
    count = 0
    for x, row in dfLinks.iterrows():
        dfLinks.loc[x]['State'] = btn[x]['text']
        count = row['Row']
        if row['realColumn'] == 'Sum Week':
            dfF.loc[count] = list([row['Row']]) + list([row['State']]) + ['Nan'] + ['Nan'] + ['Nan'] + ['Nan'] + ['Nan']
        elif row['realColumn'] == 'Sum Sat':
            dfF.loc[count]['Sum Sat'] = row['State']
        elif row['realColumn'] == 'Sum Sun':
            dfF.loc[count]['Sum Sun'] = row['State']
        elif row['realColumn'] == 'Wint Week':
            dfF.loc[count]['Wint Week'] = row['State']
        elif row['realColumn'] == 'Wint Sat':
            dfF.loc[count]['Wint Sat'] = row['State']
        elif row['realColumn'] == 'Wint Sun':
            dfF.loc[count]['Wint Sun'] = row['State']
        # dfF.to_csv('dataFrame.csv', index=False)

    return


# Begin construction of frame
summer = tk.Label(root, text="Summer")
summer.grid(column=2, row=0)
winter = tk.Label(root, text="Winter")
winter.grid(column=9, row=0)
head1 = tk.Label(root, text="Hour")
head1.grid(column=0, row=1)
head2 = tk.Label(root, text="Weekdays")
head2.grid(column=1, row=1)
head3 = tk.Label(root, text="Saturdays")
head3.grid(column=2, row=1)
head4 = tk.Label(root, text="Sunday")
head4.grid(column=3, row=1)
head5 = tk.Label(root, text="      ")
head5.grid(column=4, row=1)
head6 = tk.Label(root, text="      ")
head6.grid(column=5, row=1)
head7 = tk.Label(root, text="      ")
head7.grid(column=6, row=1)
head8 = tk.Label(root, text="      ")
head8.grid(column=7, row=1)
head9 = tk.Label(root, text="Weekdays")
head9.grid(column=8, row=1)
head10 = tk.Label(root, text="Saturdays")
head10.grid(column=9, row=1)
head11 = tk.Label(root, text="Sunday")
head11.grid(column=10, row=1)

lbl0 = tk.Label(root, text="0")
lbl0.grid(column=0, row=2)
lbl1 = tk.Label(root, text="1")
lbl1.grid(column=0, row=3)
lbl2 = tk.Label(root, text="2")
lbl2.grid(column=0, row=4)
lbl3 = tk.Label(root, text="3")
lbl3.grid(column=0, row=5)
lbl4 = tk.Label(root, text="4")
lbl4.grid(column=0, row=6)
lbl5 = tk.Label(root, text="5")
lbl5.grid(column=0, row=7)
lbl6 = tk.Label(root, text="6")
lbl6.grid(column=0, row=8)
lbl7 = tk.Label(root, text="7")
lbl7.grid(column=0, row=9)
lbl8 = tk.Label(root, text="8")
lbl8.grid(column=0, row=10)
lbl9 = tk.Label(root, text="9")
lbl9.grid(column=0, row=11)
lbl10 = tk.Label(root, text="10")
lbl10.grid(column=0, row=12)
lbl11 = tk.Label(root, text="11")
lbl11.grid(column=0, row=13)
lbl12 = tk.Label(root, text="12")
lbl12.grid(column=0, row=14)
lbl13 = tk.Label(root, text="13")
lbl13.grid(column=0, row=15)
lbl14 = tk.Label(root, text="14")
lbl14.grid(column=0, row=16)
lbl15 = tk.Label(root, text="15")
lbl15.grid(column=0, row=17)
lbl16 = tk.Label(root, text="16")
lbl16.grid(column=0, row=18)
lbl17 = tk.Label(root, text="17")
lbl17.grid(column=0, row=19)
lbl18 = tk.Label(root, text="18")
lbl18.grid(column=0, row=20)
lbl19 = tk.Label(root, text="19")
lbl19.grid(column=0, row=21)
lbl20 = tk.Label(root, text="20")
lbl20.grid(column=0, row=22)
lbl21 = tk.Label(root, text="21")
lbl21.grid(column=0, row=23)
lbl22 = tk.Label(root, text="22")
lbl22.grid(column=0, row=24)
lbl23 = tk.Label(root, text="23")
lbl23.grid(column=0, row=25)

fill1 = tk.Label(root, text="      ")
fill1.grid(column=14, row=0)

winterT = tk.Label(root, text="Winter Tariffs (N$/kWh)")
winterT.grid(column=16, row=0)

fill1 = tk.Label(root, text="      ")
fill1.grid(column=17, row=0)

summerT = tk.Label(root, text="Summer Tariffs (N$/kWh)")
summerT.grid(column=18, row=0)

peak = tk.Label(root, text="Peak")
peak.grid(column=15, row=1)
off = tk.Label(root, text="Off")
off.grid(column=15, row=3)
standard = tk.Label(root, text="Standard")
standard.grid(column=15, row=2)

winterP = tk.Entry(root)
winterP.insert(1, "2.41")
winterP.grid(column=16, row=1)
winterO = tk.Entry(root)
winterO.insert(1, "1.88")
winterO.grid(column=16, row=2)
winterS = tk.Entry(root)
winterS.insert(1, "1.42")
winterS.grid(column=16, row=3)
summerP = tk.Entry(root)
summerP.insert(1, "2.41")
summerP.grid(column=18, row=1)
summerO = tk.Entry(root)
summerO.insert(1, "1.88")
summerO.grid(column=18, row=2)
summerS = tk.Entry(root)
summerS.insert(1, "1.42")
summerS.grid(column=18, row=3)

maxDT = tk.Entry(root)
maxDT.insert(1, "138")
maxDT.grid(column=16, row=7)
maxDTL = tk.Label(root, text="Max Demand Tariff (N$/kVA)")
maxDTL.grid(column=16, row=6)

ecbLevy = tk.Entry(root)
ecbLevy.insert(1, "0.02120")
ecbLevy.grid(column=16, row=9)
ecbLevyL = tk.Label(root, text="ECB Levy (N$/kWh)")
ecbLevyL.grid(column=16, row=8)

nefLevy = tk.Entry(root)
nefLevy.insert(1, "0.016")
nefLevy.grid(column=16, row=11)
nefLevyL = tk.Label(root, text="NEF Levy (N$/kWh)")
nefLevyL.grid(column=16, row=10)

netwLevy = tk.Entry(root)
netwLevy.insert(1, "52")
netwLevy.grid(column=16, row=13)
netwLevyL = tk.Label(root, text="Network Access Charge (N$/kVA)")
netwLevyL.grid(column=16, row=12)

declaredD = tk.Entry(root)
declaredD.insert(1, "200")
declaredD.grid(column=16, row=15)
declaredDL = tk.Label(root, text="Declared Demand (kVA)")
declaredDL.grid(column=16, row=14)

netMeterTariff = tk.Entry(root)
netMeterTariff.insert(1, "1.50")
netMeterTariff.grid(column=18, row=7)
netMeterTariffL = tk.Label(root, text="Net Metering Tariff (N$/kWh)")
netMeterTariffL.grid(column=18, row=6)

stringT = 'O'
stringC = 'Green'
number = 0

# Set the buttons state
for i in range(6):  # Because you have the six rows of the winter and summer column, this is the best wat to itterate though them
    for j in range(24):  # This is the hour
        if ((j > 6 and j < 10) or (j > 16 and j < 21)) and (i == 0):
            stringT = 'P'
            stringC = 'Red'
        elif ((j == 6) or (j > 9 and j < 17) or (j == 21)) and (i == 0):
            stringT = 'S'
            stringC = 'Yellow'
        elif ((j > 6 and j < 12) or (j == 18) or (j == 19)) and (i == 1):
            stringT = 'S'
            stringC = 'Yellow'
        elif ((j > 6 and j < 10) or (j > 16 and j < 21)) and (i == 3):
            stringT = 'P'
            stringC = 'Red'
        elif ((j == 6) or (j > 9 and j < 17) or (j == 21)) and (i == 3):
            stringT = 'S'
            stringC = 'Yellow'
        elif ((j > 6 and j < 12) or (j == 18) or (j == 19)) and (i == 4):
            stringT = 'S'
            stringC = 'Yellow'
        else:
            stringT = 'O'
            stringC = 'Green'
        btn.append(tk.Button(root, text=stringT, padx=20, pady=0, fg="Black", bg=stringC,
                             command=lambda column=i, row=j, nbr=number: changeState(btn, nbr, column, row)))

        if i < 3:
            btn[number].grid(column=i + 1, row=j + 2)
        else:
            btn[number].grid(column=i + 5, row=j + 2)

        realColumn = ''
        if i == 0:
            realColumn = 'Sum Week'
        elif i == 1:
            realColumn = 'Sum Sat'
        elif i == 2:
            realColumn = 'Sum Sun'
        elif i == 3:
            realColumn = 'Wint Week'
        elif i == 4:
            realColumn = 'Wint Sat'
        elif i == 5:
            realColumn = 'Wint Sun'

        dfLinks.loc[number] = [i] + [j] + [number] + list(btn[number]['text']) + [realColumn]
        number += 1

inputBt = tk.Button(root, text="1. Select Input", bg="Grey", command=selectIn, width=16)
inputBt.grid(column=16, row=17, rowspan=2)

confirmSchedule = tk.Button(root, text="2. Confirm Schedule", bg="Grey", command=lambda: confirmSched(btn), width=16)
confirmSchedule.grid(column=16, row=19, rowspan=2)

outputBt = tk.Button(root, text="3. Generate Output", bg="Grey", command=genOutput, width=16)
outputBt.grid(column=16, row=21, rowspan=2)

quitBt = tk.Button(root, text="4. Exit", bg="Grey", command=exit, width=16)
quitBt.grid(column=16, row=23, rowspan=2)

# testBt = tk.Button(root, text="5. Text", bg="Grey", command=lambda: createExcel("This is the test title", dfTotals),
#                    width=16)
# testBt.grid(column=16, columnspan=3, row=19, rowspan=2)

root.mainloop()
